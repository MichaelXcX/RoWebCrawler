import os
import requests
import bs4
from bs4 import Comment
import openpyxl
import re
from utils import filter_domain_from_url

### Open .xlsx file
def open_workbook(file):
    try:
        wb = openpyxl.load_workbook(file)
        return wb
    except Exception as e:
        print('There was a problem: %s' % (e))

### Get into robots.txt
def get_robots(url): 
    try:
        res = requests.get(url + '/robots.txt')
        res.raise_for_status()
        robots = res.text
        return robots
    except Exception as e:
        print('There was a problem: %s' % (e))

### Extract Sitemap from robots.txt: 
def extract_sitemap(robots):
    try:
        robots = robots.split('\n')
        sitemaps = []
        for line in robots:
            if 'Sitemap' in line:
                sitemap = line.split(': ')[1]
                print(sitemap)
                sitemaps += [sitemap]
        return sitemaps
    except Exception as e:
        print('There was a problem: %s' % (e))

### Extracting links from sitemap.xml
def extract_links(sitemap):
    try:
        res = requests.get(sitemap)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        links = []
        for link in soup.find_all('loc'):
            links += [link.text]
        return links
    except Exception as e:
        print('There was a problem: %s' % (e))

### Get html page and save it in a file inside results/ folder
### Bonus: get rid of metadata and scripts from the page
def get_html(url):
    try:
        res = requests.get(url)
        res.raise_for_status()
        soup = bs4.BeautifulSoup(res.text, 'html.parser')
        
        # Removing meta tags
        metas = soup.find_all('meta')
        for meta in metas:
            meta.decompose()
        
        # Removing script tags
        scripts = soup.find_all('script')
        for script in scripts:
            script.decompose()
        
        # Removing link tags
        links = soup.find_all('link')
        for link in links:
            link.decompose()

        soupstr = str(soup.prettify());
        with open('results/' + url.split('/')[-1] + '.html', 'w') as file:
            file.write(soupstr)
    except Exception as e:
        print('There was a problem: %s' % (e))

wb = open_workbook('websites.xlsx')
sheet = wb.active
for i in range(1, sheet.max_row):
    url = sheet.cell(row=i, column=2).value
    url = filter_domain_from_url(url)
    print("----------------------------------------------------")
    print(url)
    soup = get_robots(url)
    if soup: 
        sitemaps = extract_sitemap(soup)

        if sitemaps: 
            pages = []

            ### Extracting links from sitemap.xml, eliminating .xml files   
            for sitemap in sitemaps:
                links = extract_links(sitemap)
                for link in links:
                    if '.xml' not in link:
                        pages.append(link)
        else: 
            print("No sitemap found")
            continue
    else:
        print("No robots.txt found")
        continue
    print('----------------------------------------------------')
# links = extract_links(sitemaps[0])
# for link in links:
#     if '.xml' not in link:
#         pages.append(link)

# for page in pages:
#     get_html(page)
#     print("Page saved: " + page)
#     # print(page)

# print("Done saving " + str(len(pages)) + " pages.")

